"""
Breakwater Capital Partners — Underwriting Model builder.
Generates Breakwater_Underwriting_Model.xlsx with live formulas driven off the
Inputs tab. Re-run to regenerate. Worked example: a Rio Vista / Las Olas spec home.

    python3 build_underwriting_model.py
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ---- Brand styling -------------------------------------------------------
NAVY = "0B2545"      # deep navy
STEEL = "13315C"     # steel blue
SAND = "EEF2F6"      # light panel
ACCENT = "B68D40"    # muted gold

WHITE_BOLD = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
TITLE = Font(name="Calibri", bold=True, color="FFFFFF", size=16)
SUB = Font(name="Calibri", bold=True, color=ACCENT, size=11)
LABEL = Font(name="Calibri", size=11)
LABEL_B = Font(name="Calibri", bold=True, size=11)
NOTE = Font(name="Calibri", italic=True, size=9, color="666666")
KPI = Font(name="Calibri", bold=True, size=12, color=NAVY)

fill_navy = PatternFill("solid", fgColor=NAVY)
fill_steel = PatternFill("solid", fgColor=STEEL)
fill_sand = PatternFill("solid", fgColor=SAND)
fill_gold = PatternFill("solid", fgColor=ACCENT)

thin = Side(style="thin", color="C9D2DC")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

USD = '_($* #,##0_);_($* (#,##0);_($* "-"_);_(@_)'
PCT = '0.0%'
NUM = '#,##0'

wb = Workbook()

def style_header(ws, text, ncols=6):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    c = ws.cell(row=1, column=1, value=text)
    c.font = TITLE
    c.fill = fill_navy
    c.alignment = Alignment(vertical="center", horizontal="left", indent=1)
    ws.row_dimensions[1].height = 30

def kv(ws, r, label, value, fmt=None, bold=False, panel=False):
    a = ws.cell(row=r, column=1, value=label)
    a.font = LABEL_B if bold else LABEL
    b = ws.cell(row=r, column=2, value=value)
    if fmt:
        b.number_format = fmt
    b.font = LABEL_B if bold else LABEL
    b.alignment = Alignment(horizontal="right")
    if panel:
        a.fill = fill_sand
        b.fill = fill_sand
    a.border = border
    b.border = border
    return b

# ========================================================================
# COVER
# ========================================================================
ws = wb.active
ws.title = "Cover"
ws.sheet_view.showGridLines = False
ws.column_dimensions['A'].width = 2
ws.column_dimensions['B'].width = 70
for r in range(1, 30):
    ws.row_dimensions[r].height = 18
ws.merge_cells("B2:B3")
c = ws.cell(row=2, column=2, value="BREAKWATER CAPITAL PARTNERS")
c.font = Font(name="Calibri", bold=True, size=22, color=NAVY)
ws.cell(row=5, column=2, value="Special Situations Development Platform").font = Font(size=13, color=STEEL, italic=True)
ws.cell(row=6, column=2, value="Single-Asset Underwriting Model").font = Font(size=13, bold=True, color=ACCENT)
lines = [
    "",
    "PURPOSE",
    "Underwrites one distressed luxury-residential completion opportunity from",
    "distressed basis to exit. Solve for the completion spread, the equity check,",
    "project IRR, and the partner-level waterfall — then stress-test the result.",
    "",
    "HOW TO USE",
    "1.  Edit only the yellow cells on the INPUTS tab.",
    "2.  UNDERWRITING, RETURNS, and SENSITIVITY recalculate automatically.",
    "3.  Confirm the deal clears the thresholds (see methodology doc).",
    "",
    "DECISION THRESHOLDS (must all hold to proceed)",
    "   •  Completion spread  ≥ 20% base / target 25%+",
    "   •  Downside case does NOT lose capital",
    "   •  Project gross IRR  ≥ 20%   ·   Equity multiple ≥ 1.35x",
    "   •  Contingency carried ≥ 10%",
    "",
    "Worked example loaded: illustrative Rio Vista / Las Olas waterfront spec home.",
]
r = 8
for ln in lines:
    cell = ws.cell(row=r, column=2, value=ln)
    if ln in ("PURPOSE", "HOW TO USE", "DECISION THRESHOLDS (must all hold to proceed)"):
        cell.font = Font(bold=True, size=11, color=NAVY)
    elif ln.startswith("Worked example"):
        cell.font = NOTE
    else:
        cell.font = Font(size=10, color="333333")
    r += 1
ws.cell(row=r+1, column=2,
        value="Illustrative only. Not investment advice; not a guarantee of results.").font = NOTE

# ========================================================================
# INPUTS  (only tab the user edits)
# ========================================================================
wi = wb.create_sheet("Inputs")
wi.sheet_view.showGridLines = False
wi.column_dimensions['A'].width = 40
wi.column_dimensions['B'].width = 18
wi.column_dimensions['C'].width = 40
style_header(wi, "INPUTS  —  edit the yellow cells only", 3)
wi.cell(row=2, column=3, value="Source / discipline").font = SUB

inputs = [
    # row, label, value, fmt, note
    ("Finished (exit) value",        7_500_000, USD, "≥3 finished comps; brokerage-validated; time haircut"),
    ("Acquisition cost (control)",   3_700_000, USD, "Solved from screen + max-basis, not the ask"),
    ("Acquisition closing %",        0.02,      PCT, "Title, legal, transfer; in-house where possible"),
    ("Cost to complete (hard+soft)", 1_400_000, USD, "Independent estimate; scope-verified"),
    ("Contingency %",                0.12,      PCT, "10% base / 12–15% complex or stale budget"),
    ("Hold period (months)",         12,        NUM, "Construction + listing + closing lag; honest"),
    ("Monthly carry",               15_000,     USD, "Taxes + insurance + utilities/security + admin"),
    ("Disposition cost %",           0.05,      PCT, "Commission + closing + concessions"),
    ("Loan amount (debt)",          4_800_000,  USD, "Per term sheet; preserve refi optionality"),
    ("Loan rate (annual)",           0.10,      PCT, "Acquisition / completion financing"),
    ("Origination / fees %",         0.015,     PCT, "Points + lender fees"),
    ("Preferred return (annual)",    0.09,      PCT, "Hurdle to capital partners before promote"),
    ("Promote (GP share over pref)", 0.20,      PCT, "GP share of residual profit"),
]
# map label -> input row for formula references
row0 = 4
ref = {}
for i, (label, val, fmt, note) in enumerate(inputs):
    r = row0 + i
    a = wi.cell(row=r, column=1, value=label); a.font = LABEL_B; a.border = border
    b = wi.cell(row=r, column=2, value=val); b.number_format = fmt
    b.font = Font(bold=True, color=NAVY); b.fill = fill_gold
    b.alignment = Alignment(horizontal="right"); b.border = border
    n = wi.cell(row=r, column=3, value=note); n.font = NOTE; n.border = border
    ref[label] = f"Inputs!$B${r}"

# convenient refs
FV   = ref["Finished (exit) value"]
ACQ  = ref["Acquisition cost (control)"]
ACQC = ref["Acquisition closing %"]
CTC  = ref["Cost to complete (hard+soft)"]
CONT = ref["Contingency %"]
HOLD = ref["Hold period (months)"]
CARRY= ref["Monthly carry"]
DISP = ref["Disposition cost %"]
LOAN = ref["Loan amount (debt)"]
RATE = ref["Loan rate (annual)"]
ORIG = ref["Origination / fees %"]
PREF = ref["Preferred return (annual)"]
PROMO= ref["Promote (GP share over pref)"]

# ========================================================================
# UNDERWRITING
# ========================================================================
wu = wb.create_sheet("Underwriting")
wu.sheet_view.showGridLines = False
wu.column_dimensions['A'].width = 42
wu.column_dimensions['B'].width = 18
style_header(wu, "UNDERWRITING  —  all-in basis, spread & project return", 2)

r = 3
wu.cell(row=r, column=1, value="PROJECT COSTS (asset-level, unlevered)").font = SUB
r += 1
kv(wu, r, "Acquisition cost (control)", f"={ACQ}", USD); ACQ_r = r; r += 1
kv(wu, r, "Acquisition closing costs", f"={ACQ}*{ACQC}", USD); ACQC_r = r; r += 1
kv(wu, r, "Cost to complete (hard+soft)", f"={CTC}", USD); CTC_r = r; r += 1
kv(wu, r, "Contingency", f"={CTC}*{CONT}", USD); CONTr = r; r += 1
kv(wu, r, "Carry costs over hold", f"={CARRY}*{HOLD}", USD); CARRYr = r; r += 1
kv(wu, r, "Disposition costs", f"={FV}*{DISP}", USD); DISPr = r; r += 1
kv(wu, r, "ALL-IN BASIS (asset-level)",
   f"=SUM(B{ACQ_r}:B{DISPr})", USD, bold=True, panel=True); BASIS_r = r; r += 2

wu.cell(row=r, column=1, value="COMPLETION SPREAD").font = SUB
r += 1
kv(wu, r, "Finished (exit) value", f"={FV}", USD); r += 1
kv(wu, r, "Completion spread ($)", f"={FV}-B{BASIS_r}", USD, bold=True); SPRD_r = r; r += 1
sp = kv(wu, r, "Completion spread (%)", f"=B{SPRD_r}/B{BASIS_r}", PCT, bold=True, panel=True)
SPRDP_r = r; r += 2

wu.cell(row=r, column=1, value="FINANCING").font = SUB
r += 1
kv(wu, r, "Interest (IO over hold)", f"={LOAN}*{RATE}*{HOLD}/12", USD); INT_r = r; r += 1
kv(wu, r, "Origination / fees", f"={LOAN}*{ORIG}", USD); ORIGr = r; r += 1
kv(wu, r, "Total financing cost", f"=B{INT_r}+B{ORIGr}", USD, bold=True); FIN_r = r; r += 2

wu.cell(row=r, column=1, value="SOURCES & USES / EQUITY").font = SUB
r += 1
kv(wu, r, "Total uses (basis + financing)", f"=B{BASIS_r}+B{FIN_r}", USD); USES_r = r; r += 1
kv(wu, r, "Less: loan (debt)", f"=-{LOAN}", USD); r += 1
kv(wu, r, "EQUITY REQUIRED", f"=B{USES_r}-{LOAN}", USD, bold=True, panel=True); EQ_r = r; r += 2

wu.cell(row=r, column=1, value="PROJECT RETURN (to equity)").font = SUB
r += 1
kv(wu, r, "Net sale proceeds (FV − disposition)", f"={FV}-{FV}*{DISP}", USD); NET_r = r; r += 1
kv(wu, r, "Equity proceeds at exit (net − loan)", f"=B{NET_r}-{LOAN}", USD); EXIT_r = r; r += 1
kv(wu, r, "PROJECT PROFIT", f"=B{EXIT_r}-B{EQ_r}", USD, bold=True); PROF_r = r; r += 1
em = kv(wu, r, "Equity multiple (gross)", f"=B{EXIT_r}/B{EQ_r}", '0.00"x"', bold=True, panel=True)
EM_r = r; r += 1
ir = kv(wu, r, "Annualized return (approx IRR)",
        f"=(B{EXIT_r}/B{EQ_r})^(12/{HOLD})-1", PCT, bold=True, panel=True)
IRR_r = r; r += 2

# threshold check flags
wu.cell(row=r, column=1, value="THRESHOLD CHECK").font = SUB
r += 1
kv(wu, r, "Spread ≥ 20% ?", f'=IF(B{SPRDP_r}>=0.2,"PASS","REVIEW")', None);
wu.cell(row=r, column=2).alignment = Alignment(horizontal="right"); r += 1
kv(wu, r, "Equity multiple ≥ 1.35x ?", f'=IF(B{EM_r}>=1.35,"PASS","REVIEW")', None)
wu.cell(row=r, column=2).alignment = Alignment(horizontal="right"); r += 1
kv(wu, r, "IRR ≥ 20% ?", f'=IF(B{IRR_r}>=0.2,"PASS","REVIEW")', None)
wu.cell(row=r, column=2).alignment = Alignment(horizontal="right"); r += 1

# expose key cells for other tabs
U_EQ = f"Underwriting!$B${EQ_r}"
U_EXIT = f"Underwriting!$B${EXIT_r}"
U_PROF = f"Underwriting!$B${PROF_r}"

# ========================================================================
# RETURNS — waterfall
# ========================================================================
wr = wb.create_sheet("Returns")
wr.sheet_view.showGridLines = False
wr.column_dimensions['A'].width = 46
wr.column_dimensions['B'].width = 18
style_header(wr, "RETURNS  —  partner waterfall (deal-by-deal)", 2)
r = 3
wr.cell(row=r, column=1, value="DISTRIBUTION WATERFALL").font = SUB
r += 1
kv(wr, r, "Invested equity (capital partners)", f"={U_EQ}", USD); INVr = r; r += 1
kv(wr, r, "Total distributable (equity proceeds)", f"={U_EXIT}", USD); DISTr = r; r += 1
kv(wr, r, "Preferred return accrued", f"={U_EQ}*{PREF}*{HOLD}/12", USD); PREFr = r; r += 1
kv(wr, r, "Project profit", f"={U_PROF}", USD); PJr = r; r += 2

wr.cell(row=r, column=1, value="TIERED SPLIT").font = SUB
r += 1
kv(wr, r, "1. Return of capital → LP", f"=MIN(B{DISTr},B{INVr})", USD); T1 = r; r += 1
kv(wr, r, "2. Preferred return → LP", f"=MIN(B{PJr},B{PREFr})", USD); T2 = r; r += 1
kv(wr, r, "   Residual after pref", f"=MAX(B{PJr}-B{PREFr},0)", USD); RESr = r; r += 1
kv(wr, r, "3. Promote → GP", f"=B{RESr}*{PROMO}", USD); GPr = r; r += 1
kv(wr, r, "4. Residual split → LP", f"=B{RESr}*(1-{PROMO})", USD); LPRr = r; r += 2

wr.cell(row=r, column=1, value="PARTNER OUTCOMES").font = SUB
r += 1
kv(wr, r, "LP total distribution", f"=B{T1}+B{T2}+B{LPRr}", USD, bold=True); LPT = r; r += 1
kv(wr, r, "LP profit", f"=B{T2}+B{LPRr}", USD, bold=True); r += 1
kv(wr, r, "LP equity multiple", f"=B{LPT}/B{INVr}", '0.00"x"', bold=True, panel=True); LPMr = r; r += 1
kv(wr, r, "LP annualized return", f"=(B{LPT}/B{INVr})^(12/{HOLD})-1", PCT, bold=True, panel=True); r += 1
kv(wr, r, "GP profit (promote)", f"=B{GPr}", USD, bold=True, panel=True); r += 1
wr.cell(row=r+1, column=1,
        value="GP also co-invests alongside LPs on every deal (not modeled here).").font = NOTE

# ========================================================================
# SENSITIVITY — 2-way grids (recomputed inline so they stay live)
# ========================================================================
ws2 = wb.create_sheet("Sensitivity")
ws2.sheet_view.showGridLines = False
style_header(ws2, "SENSITIVITY  —  finished value × cost to complete", 8)
ws2.column_dimensions['A'].width = 26
for col in "BCDEF":
    ws2.column_dimensions[col].width = 15

fv_mults  = [0.90, 0.95, 1.00, 1.05, 1.10]      # columns
ctc_mults = [0.90, 1.00, 1.10, 1.20]            # rows

def profit_formula(fvm, ctcm):
    """Asset-level project profit with FV and CTC scaled. Loan/financing fixed."""
    fv = f"({FV}*{fvm})"
    ctc = f"({CTC}*{ctcm})"
    basis = (f"{ACQ}+{ACQ}*{ACQC}+{ctc}+{ctc}*{CONT}+{CARRY}*{HOLD}+{fv}*{DISP}")
    fin = f"{LOAN}*{RATE}*{HOLD}/12+{LOAN}*{ORIG}"
    return f"={fv}-({basis})-({fin})"

def spread_formula(fvm, ctcm):
    fv = f"({FV}*{fvm})"
    ctc = f"({CTC}*{ctcm})"
    basis = (f"{ACQ}+{ACQ}*{ACQC}+{ctc}+{ctc}*{CONT}+{CARRY}*{HOLD}+{fv}*{DISP}")
    return f"=({fv}-({basis}))/({basis})"

def grid(ws, top, title, formula_fn, fmt):
    ws.cell(row=top, column=1, value=title).font = SUB
    ws.cell(row=top+1, column=2, value="Finished value →").font = NOTE
    ws.cell(row=top+2, column=1, value="Cost to complete ↓").font = LABEL_B
    # column headers
    for j, fvm in enumerate(fv_mults):
        c = ws.cell(row=top+2, column=2+j, value=fvm-1)
        c.number_format = '+0%;-0%'
        c.font = WHITE_BOLD; c.fill = fill_steel
        c.alignment = Alignment(horizontal="center"); c.border = border
    a0 = ws.cell(row=top+2, column=1); a0.fill = fill_steel; a0.border = border
    # rows
    for i, ctcm in enumerate(ctc_mults):
        rr = top+3+i
        h = ws.cell(row=rr, column=1, value=ctcm-1)
        h.number_format = '+0%;-0%'; h.font = WHITE_BOLD; h.fill = fill_steel
        h.alignment = Alignment(horizontal="center"); h.border = border
        for j, fvm in enumerate(fv_mults):
            c = ws.cell(row=rr, column=2+j, value=formula_fn(fvm, ctcm))
            c.number_format = fmt
            c.alignment = Alignment(horizontal="right")
            c.border = border
            if fvm == 1.0 and ctcm == 1.0:
                c.fill = fill_gold; c.font = LABEL_B
    return top+3+len(ctc_mults)

nxt = grid(ws2, 3, "PROJECT PROFIT (asset-level)", profit_formula, USD)
grid(ws2, nxt+2, "COMPLETION SPREAD %", spread_formula, PCT)
ws2.cell(row=nxt+2+3+len(ctc_mults)+1, column=1,
         value="Gold cell = base case. Confirm the downside corner (−10% value / +20% cost) "
               "does not destroy the spread.").font = NOTE

# ---- order & save --------------------------------------------------------
wb.move_sheet("Cover", -wb.index(wb["Cover"]))
out = "Breakwater_Underwriting_Model.xlsx"
wb.save(out)
print("Saved", out, "with tabs:", wb.sheetnames)
